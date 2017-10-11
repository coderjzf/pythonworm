# -*- coding: utf-8 -*-


# 进入80s首页,点击电影选项,在点击好评选项,进入按照电影豆瓣评分排名
# 按照豆瓣评分排名首页为http://www.80s.tw/movie/list/----g
# 第2页为http://www.80s.tw/movie/list/----g-p2
# 所以可以根据p1,p2,p3...来进行分页


# 利用requests模块模拟请求
import requests

# 利用BeautifulSoup解析html
from bs4 import BeautifulSoup

# 利用openpyxl将数据写入excel
from openpyxl import Workbook

# 利用time控制程序休眠
import time

# 全局变量row,表示excel中表单的行
row = 2


# 获取单个页面
def get_one_page(url):
    try:
        # 模拟get请求
        response = requests.get(url, timeout=5)
        if response.status_code == 200:
            return response.text
    except requests.ConnectionError as e:
        print(e.reason)


# 解析网页
def parse_one_page(html):
    soup = BeautifulSoup(html, "html.parser")
    # 网站首页
    base_url = "http://www.80s.tw"
    # 电影列表
    items = soup.find_all("ul")[9]

    # 每部电影的信息都包含在一个li标签内
    for item in items.find_all("li"):
        # 电影名
        name = item.find("img").attrs["alt"]
        # 评分
        score = item.find(attrs={"class": "poster_score"}).string
        # 详情页
        # 每个li标签的第1个a标签的href为电影详情页的相对url,加上网站首页(根目录)为完整的url
        src = base_url + item.find("a").attrs["href"]
        # 短评
        quote = item.find(attrs={"class": "tip"}).string.strip()

        # 必须使用global,否则每次循环都会使用row的初始值
        global row

        # 将电影信息写入表单
        sh.cell(row=row, column=1).value = name
        sh.cell(row=row, column=2).value = score
        sh.cell(row=row, column=3).value = src
        sh.cell(row=row, column=4).value = quote

        print(name, score, src, quote)

        row = row + 1


# 主方法,固定写法
if __name__ == '__main__':

    # 创建一个工作簿
    wb = Workbook()
    # 创建一个表单
    wb.create_sheet("80s电影Top1000", 0)
    # 向表单中写入各项名称
    sh = wb.active
    sh.cell(row=1, column=1).value = "电影名"
    sh.cell(row=1, column=2).value = "评分"
    sh.cell(row=1, column=3).value = "详情页"
    sh.cell(row=1, column=4).value = "短评"

    # 根据page进行分页
    for page in range(1, 40):
        # 要抓取的url
        url = "http://www.80s.tw/movie/list/----g-p" + str(page)

        # 获取单个网页
        html = get_one_page(url)

        # 解析网页并将,电影的各项信息写入excel表单
        parse_one_page(html)

        # 休眠2秒
        time.sleep(2)

    # 保存excel文件
    wb.save("80s电影Top1000.xlsx")
